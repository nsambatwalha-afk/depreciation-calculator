import io

import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date
from datetime import timedelta

st.set_page_config(page_title="Depreciation Calculator", layout="wide")

# ----------------------------------------------------
# EXPECTED FAR FORMAT
# ----------------------------------------------------
EXPECTED_COLUMNS = [
    "Asset ID",
    "Asset Name",
    "Asset Category",
    "Cost",
    "Rate",
    "Acquisition Date"
]

EXPECTED_DISPOSAL_COLUMNS = EXPECTED_COLUMNS + ["Disposal Date"]


# ----------------------------------------------------
# STRAIGHT LINE CALCULATOR
# ----------------------------------------------------
def calculator(cost, acquisition_date, period_start, period_end, rate):

    ul_days = int((1 / rate) * 365)
    eol = acquisition_date + timedelta(days=ul_days)
    dep_daily = (rate * cost) / 365

    if period_start >= eol:
        return 0.0

    dep_start = max(acquisition_date, period_start)
    dep_end = min(period_end, eol)

    dep_days = (dep_end - dep_start).days + 1

    if dep_days <= 0:
        return 0.0

    return dep_daily * dep_days


# ----------------------------------------------------
# REDUCING BALANCE CALCULATOR
# ----------------------------------------------------
def calculator2(cost, acquisition_date, period_start, period_end, rate):

    ul_days = int((1 / rate) * 365)
    eol = acquisition_date + timedelta(days=ul_days)

    if period_start >= eol:
        return 0.0

    dep_start = max(acquisition_date, period_start)
    dep_end = min(period_end, eol)

    dep_days = (dep_end - dep_start).days + 1

    if dep_days <= 0:
        return 0.0

    days_before_period = (dep_start - acquisition_date).days
    years_elapsed = days_before_period / 365

    book_value = cost * ((1 - rate) ** years_elapsed)

    dep_daily = (rate * book_value) / 365

    return dep_daily * dep_days


# ----------------------------------------------------
# ACCUMULATED DEPRECIATION — STRAIGHT LINE
# ----------------------------------------------------
def calc_accumulated_sl(cost, acquisition_date, end_date, rate):
    """Calculate accumulated depreciation from acquisition_date to end_date."""

    if rate <= 0:
        return 0.0

    ul_days = int((1 / rate) * 365)
    eol = acquisition_date + timedelta(days=ul_days)
    dep_daily = (rate * cost) / 365

    if acquisition_date >= end_date:
        return 0.0

    actual_end = min(end_date, eol)
    acc_days = (actual_end - acquisition_date).days

    if acc_days <= 0:
        return 0.0

    return dep_daily * acc_days


# ----------------------------------------------------
# ACCUMULATED DEPRECIATION — REDUCING BALANCE
# ----------------------------------------------------
def calc_accumulated_rb(cost, acquisition_date, end_date, rate):
    """Calculate accumulated depreciation from acquisition_date to end_date."""

    if rate <= 0:
        return 0.0

    ul_days = int((1 / rate) * 365)
    eol = acquisition_date + timedelta(days=ul_days)

    if acquisition_date >= end_date:
        return 0.0

    actual_end = min(end_date, eol)
    years_elapsed = (actual_end - acquisition_date).days / 365

    book_value = cost * ((1 - rate) ** years_elapsed)
    accumulated_dep = cost - book_value

    return max(0.0, accumulated_dep)


# ----------------------------------------------------
# DEPRECIATION ENGINE
# ----------------------------------------------------
def calculate_depreciation(far_df, period_start, period_end, method, disposal_df=None):

    results_list = []
    category_totals = {}
    accumulated_by_category = {}

    # Build a lookup of disposed assets: {asset_id: disposal_date}
    disposed_assets = {}
    if disposal_df is not None:
        for _, row in disposal_df.iterrows():
            disposed_assets[str(row["Asset ID"])] = row["Disposal Date"].date()

    for _, row in far_df.iterrows():

        asset_id = row["Asset ID"]
        asset_name = row["Asset Name"]
        category = row["Asset Category"]
        cost = row["Cost"]
        rate = row["Rate"]

        acquisition_date = row["Acquisition Date"].date()

        disposal_date = disposed_assets.get(str(asset_id))
        is_disposed = disposal_date is not None

        # --------------------------------------------------
        # PERIOD DEPRECIATION CHARGE
        # --------------------------------------------------
        if is_disposed and disposal_date < period_start:
            # Disposed before this period — no charge
            depreciation = 0.0
        elif is_disposed and disposal_date <= period_end:
            # Disposed during this period — charge only up to disposal date
            if method == "Straight Line":
                depreciation = calculator(
                    cost, acquisition_date, period_start, disposal_date, rate
                )
            else:
                depreciation = calculator2(
                    cost, acquisition_date, period_start, disposal_date, rate
                )
        else:
            # Active asset — full period charge
            if method == "Straight Line":
                depreciation = calculator(
                    cost, acquisition_date, period_start, period_end, rate
                )
            else:
                depreciation = calculator2(
                    cost, acquisition_date, period_start, period_end, rate
                )

        if depreciation is None:
            depreciation = 0.0

        # --------------------------------------------------
        # ACCUMULATED DEPRECIATION
        # Disposed assets are removed from the books, so their
        # accumulated depreciation is excluded from the closing balance.
        # --------------------------------------------------
        if is_disposed:
            accumulated_dep = 0.0
        else:
            if method == "Straight Line":
                accumulated_dep = calc_accumulated_sl(
                    cost, acquisition_date, period_end, rate
                )
            else:
                accumulated_dep = calc_accumulated_rb(
                    cost, acquisition_date, period_end, rate
                )

        status = "Disposed" if is_disposed else "Active"

        results_list.append({
            "Asset ID": asset_id,
            "Asset Name": asset_name,
            "Asset Category": category,
            "Cost": cost,
            "Status": status,
            "Depreciation Charge": depreciation,
            "Accumulated Depreciation": accumulated_dep
        })

        if category not in category_totals:
            category_totals[category] = 0
            accumulated_by_category[category] = 0

        category_totals[category] += depreciation
        accumulated_by_category[category] += accumulated_dep

    results = pd.DataFrame(results_list)

    summary_data = [
        {
            "Asset Category": cat,
            "Total Depreciation": category_totals[cat],
            "Accumulated Depreciation": accumulated_by_category[cat]
        }
        for cat in category_totals
    ]

    summary = pd.DataFrame(summary_data)

    return results, summary


# ----------------------------------------------------
# FIXED ASSET SCHEDULE (NOTES TO FINANCIAL STATEMENTS)
# ----------------------------------------------------
def generate_fixed_asset_schedule(far_df, disposal_df, period_start, period_end, method):
    """
    Generate the fixed asset schedule as normally prepared in the notes to the
    financial statements, covering Cost, Accumulated Depreciation and Net Book
    Value by asset category.  Returns the schedule as Excel (.xlsx) bytes.
    """

    # Build a lookup of disposed assets: {asset_id: disposal_date}
    disposed_assets = {}
    if disposal_df is not None:
        for _, row in disposal_df.iterrows():
            disposed_assets[str(row["Asset ID"])] = row["Disposal Date"].date()

    # Preserve the order categories appear in the FAR
    categories = list(far_df["Asset Category"].unique())

    # Per-category schedule data
    sched = {
        cat: {
            "cost_opening": 0.0,
            "cost_additions": 0.0,
            "cost_disposals": 0.0,
            "accum_dep_opening": 0.0,
            "dep_charge": 0.0,
            "accum_dep_on_disposal": 0.0,
        }
        for cat in categories
    }

    for _, row in far_df.iterrows():
        asset_id = str(row["Asset ID"])
        category = row["Asset Category"]
        cost = row["Cost"]
        rate = row["Rate"]
        acquisition_date = row["Acquisition Date"].date()

        disposal_date = disposed_assets.get(asset_id)
        disposed_before_period = disposal_date is not None and disposal_date < period_start
        disposed_during_period = (
            disposal_date is not None
            and period_start <= disposal_date <= period_end
        )
        acquired_before_period = acquisition_date < period_start
        acquired_during_period = period_start <= acquisition_date <= period_end

        # ------ COST ------
        if acquired_before_period and not disposed_before_period:
            sched[category]["cost_opening"] += cost
        if acquired_during_period:
            sched[category]["cost_additions"] += cost
        if disposed_during_period:
            sched[category]["cost_disposals"] += cost

        # ------ ACCUMULATED DEPRECIATION: OPENING ------
        if acquired_before_period and not disposed_before_period:
            if method == "Straight Line":
                accum_open = calc_accumulated_sl(cost, acquisition_date, period_start, rate)
            else:
                accum_open = calc_accumulated_rb(cost, acquisition_date, period_start, rate)
            sched[category]["accum_dep_opening"] += accum_open

        # ------ DEPRECIATION CHARGE FOR THE PERIOD ------
        if not disposed_before_period:
            if disposed_during_period:
                if method == "Straight Line":
                    dep = calculator(cost, acquisition_date, period_start, disposal_date, rate)
                else:
                    dep = calculator2(cost, acquisition_date, period_start, disposal_date, rate)
            else:
                if method == "Straight Line":
                    dep = calculator(cost, acquisition_date, period_start, period_end, rate)
                else:
                    dep = calculator2(cost, acquisition_date, period_start, period_end, rate)
            sched[category]["dep_charge"] += dep or 0.0

        # ------ ACCUMULATED DEPRECIATION REVERSED ON DISPOSAL ------
        if disposed_during_period:
            if method == "Straight Line":
                accum_disp = calc_accumulated_sl(cost, acquisition_date, disposal_date, rate)
            else:
                accum_disp = calc_accumulated_rb(cost, acquisition_date, disposal_date, rate)
            sched[category]["accum_dep_on_disposal"] += accum_disp

    # Derive closing balances and net book values
    for cat in categories:
        d = sched[cat]
        d["cost_closing"] = d["cost_opening"] + d["cost_additions"] - d["cost_disposals"]
        d["accum_dep_closing"] = (
            d["accum_dep_opening"] + d["dep_charge"] - d["accum_dep_on_disposal"]
        )
        d["nbv_opening"] = d["cost_opening"] - d["accum_dep_opening"]
        d["nbv_closing"] = d["cost_closing"] - d["accum_dep_closing"]

    # ------ BUILD EXCEL WORKBOOK ------
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Fixed Asset Schedule"

    num_cats = len(categories)
    last_col = num_cats + 2  # label col + one col per category + Total col

    # Styles
    title_font = Font(bold=True, size=14)
    section_font = Font(bold=True)
    closing_font = Font(bold=True)
    center_align = Alignment(horizontal="center")
    right_align = Alignment(horizontal="right")
    thin_side = Side(style="thin")
    thick_side = Side(style="medium")
    num_fmt = "#,##0.00;(#,##0.00)"

    # ----- Row 1: Title -----
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
    cell = ws.cell(row=1, column=1, value="PROPERTY, PLANT AND EQUIPMENT")
    cell.font = title_font
    cell.alignment = center_align

    # ----- Row 2: Period subtitle -----
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=last_col)
    cell = ws.cell(
        row=2, column=1,
        value=(
            f"For the period {period_start.strftime('%d %B %Y')} "
            f"to {period_end.strftime('%d %B %Y')} "
            f"({method} method)"
        ),
    )
    cell.alignment = center_align

    # ----- Row 3: blank -----

    # ----- Row 4: Column headers -----
    ws.cell(row=4, column=1, value="")
    for i, cat in enumerate(categories):
        cell = ws.cell(row=4, column=2 + i, value=cat)
        cell.font = section_font
        cell.alignment = center_align
    cell = ws.cell(row=4, column=last_col, value="Total")
    cell.font = section_font
    cell.alignment = center_align
    for c in range(1, last_col + 1):
        ws.cell(row=4, column=c).border = Border(bottom=thin_side)

    # Helper: write a data row
    def write_data_row(row_num, label, field_key, negate=False, bold_row=False):
        label_cell = ws.cell(row=row_num, column=1, value=label)
        if bold_row:
            label_cell.font = closing_font
            label_cell.border = Border(top=thin_side, bottom=thick_side)

        row_total = 0.0
        for i, cat in enumerate(categories):
            val = sched[cat][field_key]
            if negate:
                val = -val
            row_total += val
            cell = ws.cell(row=row_num, column=2 + i, value=val)
            cell.number_format = num_fmt
            cell.alignment = right_align
            if bold_row:
                cell.font = closing_font
                cell.border = Border(top=thin_side, bottom=thick_side)

        total_cell = ws.cell(row=row_num, column=last_col, value=row_total)
        total_cell.number_format = num_fmt
        total_cell.alignment = right_align
        if bold_row:
            total_cell.font = closing_font
            total_cell.border = Border(top=thin_side, bottom=thick_side)

    # ----- COST section -----
    ws.cell(row=5, column=1, value="COST").font = section_font
    write_data_row(6, "Opening balance", "cost_opening")
    write_data_row(7, "Additions", "cost_additions")
    write_data_row(8, "Disposals", "cost_disposals", negate=True)
    write_data_row(9, "Closing balance", "cost_closing", bold_row=True)

    # ----- ACCUMULATED DEPRECIATION section -----
    ws.cell(row=11, column=1, value="ACCUMULATED DEPRECIATION").font = section_font
    write_data_row(12, "Opening balance", "accum_dep_opening")
    write_data_row(13, "Charge for the period", "dep_charge")
    write_data_row(14, "On disposals", "accum_dep_on_disposal", negate=True)
    write_data_row(15, "Closing balance", "accum_dep_closing", bold_row=True)

    # ----- NET BOOK VALUE section -----
    ws.cell(row=17, column=1, value="NET BOOK VALUE").font = section_font
    write_data_row(18, "Opening balance", "nbv_opening")
    write_data_row(19, "Closing balance", "nbv_closing", bold_row=True)

    # Column widths
    ws.column_dimensions[get_column_letter(1)].width = 30
    for c in range(2, last_col + 1):
        ws.column_dimensions[get_column_letter(c)].width = 18

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


# ----------------------------------------------------
# UI
# ----------------------------------------------------
st.title("Asset Depreciation Calculator")

st.write(
    "Upload the Fixed Asset Register (FAR) in the format of the template to calculate depreciation."
)


# template download
with open("template.xlsx", "rb") as file:
    template_bytes = file.read()

st.download_button(
    label="Download FAR Template",
    data=template_bytes,
    file_name="FAR_template.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)

# disposal template download
with open("disposal_template.xlsx", "rb") as file:
    disposal_template_bytes = file.read()

st.download_button(
    label="Download Disposal Template",
    data=disposal_template_bytes,
    file_name="disposal_template.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)


# upload FAR
uploaded_file = st.file_uploader(
    "Upload FAR file in the format of the template",
    type=["xlsx"]
)


if uploaded_file:

    far_df = pd.read_excel(uploaded_file)
    far_df = far_df.dropna(how="all")

    # ----------------------------------------------------
    # VALIDATE COLUMN FORMAT
    # ----------------------------------------------------
    uploaded_columns = list(far_df.columns)

    missing_columns = set(EXPECTED_COLUMNS) - set(uploaded_columns)
    extra_columns = set(uploaded_columns) - set(EXPECTED_COLUMNS)

    if missing_columns or extra_columns:

        st.error(
            "The uploaded FAR does not match the required template format. "
            "Please download and use the provided template."
        )

        if missing_columns:
            st.write("Missing columns:", list(missing_columns))

        if extra_columns:
            st.write("Unexpected columns:", list(extra_columns))

        st.stop()

    # ----------------------------------------------------
    # DATE PARSING
    # ----------------------------------------------------
    far_df["Acquisition Date"] = pd.to_datetime(
        far_df["Acquisition Date"],
        dayfirst=True
    )

    st.subheader("Uploaded FAR")
    st.dataframe(far_df, use_container_width=True)

    # ----------------------------------------------------
    # OPTIONAL DISPOSAL FILE
    # ----------------------------------------------------
    st.subheader("Disposed Assets (Optional)")
    st.write(
        "Upload the filled disposal template to exclude disposed assets from "
        "accumulated depreciation and adjust period depreciation charges accordingly."
    )

    disposal_file = st.file_uploader(
        "Upload Disposal file in the format of the disposal template",
        type=["xlsx"],
        key="disposal_uploader"
    )

    disposal_df = None

    if disposal_file:

        disposal_raw = pd.read_excel(disposal_file)
        disposal_raw = disposal_raw.dropna(how="all")

        disposal_columns = list(disposal_raw.columns)
        missing_disp = set(EXPECTED_DISPOSAL_COLUMNS) - set(disposal_columns)
        extra_disp = set(disposal_columns) - set(EXPECTED_DISPOSAL_COLUMNS)

        if missing_disp or extra_disp:

            st.error(
                "The uploaded disposal file does not match the required disposal "
                "template format. Please download and use the provided disposal template."
            )

            if missing_disp:
                st.write("Missing columns:", list(missing_disp))

            if extra_disp:
                st.write("Unexpected columns:", list(extra_disp))

        else:

            disposal_raw["Acquisition Date"] = pd.to_datetime(
                disposal_raw["Acquisition Date"],
                dayfirst=True
            )
            disposal_raw["Disposal Date"] = pd.to_datetime(
                disposal_raw["Disposal Date"],
                dayfirst=True
            )

            disposal_df = disposal_raw

            st.success(f"{len(disposal_df)} disposed asset(s) loaded.")
            st.dataframe(disposal_df, use_container_width=True)

    # depreciation period
    st.subheader("Select Depreciation Period")

    col1, col2 = st.columns(2)

    with col1:
        period_start = st.date_input("Period Start")

    with col2:
        period_end = st.date_input("Period End")

    # method dropdown
    method = st.selectbox(
        "Depreciation Method",
        ["Straight Line", "Reducing Balance"]
    )

    # run depreciation
    if st.button("Run Depreciation"):

        results, summary = calculate_depreciation(
            far_df,
            period_start,
            period_end,
            method,
            disposal_df=disposal_df
        )

        st.success("Depreciation calculation complete.")

        st.subheader("Depreciation Summary")
        st.dataframe(summary, use_container_width=True)

        st.subheader("Full FAR with Depreciation")
        st.dataframe(results, use_container_width=True)

        summary_csv = summary.to_csv(index=False)
        results_csv = results.to_csv(index=False)

        col1, col2 = st.columns(2)

        with col1:
            st.download_button(
                label="Download Summary",
                data=summary_csv,
                file_name="depreciation_summary.csv",
                mime="text/csv"
            )

        with col2:
            st.download_button(
                label="Download Full FAR with Depreciation",
                data=results_csv,
                file_name="far_with_depreciation.csv",
                mime="text/csv"
            )

        st.subheader("Fixed Asset Schedule")
        st.write(
            "Download the fixed asset schedule as prepared in the notes to the financial statements."
        )
        schedule_bytes = generate_fixed_asset_schedule(
            far_df, disposal_df, period_start, period_end, method
        )
        st.download_button(
            label="Download Fixed Asset Schedule",
            data=schedule_bytes,
            file_name="fixed_asset_schedule.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
